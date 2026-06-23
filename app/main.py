import base64
import json
import os
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import httpx
import socketio
from fastapi import Depends, FastAPI, HTTPException, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware

from .config import (
    CHAT_KB_XLSX_PATH,
    DIST_DIR,
    FLAPPY_DIR,
    NODE_ENV,
    PUBLIC_DIR,
    SESSION_SECRET,
    SQLITE_PATH,
    UNIQ_NVIDIA_API_BASE,
    UNIQ_NVIDIA_API_KEY,
    UNIQ_NVIDIA_CHAT_MODEL,
    WEB_ORIGIN,
)
from .db import bcrypt_check, bcrypt_hash, db, execute, row, rows
from .services import (
    VALID_STATUSES,
    advisor_scope,
    booking_callable_now,
    compute_estimated_minutes,
    count_words,
    csv_response,
    format_queue_number,
    get_live_queue,
    get_queue_session,
    insert_visit_log,
    minutes_between,
    next_queue_number,
    parse_dt,
    parse_school_scopes,
    parse_study_duration,
    parse_ymd,
    pick_route_advisor_id,
    recompute_route_owners,
    registration_open_for_student,
    ticket_matches_scope,
    today_sql,
)


ALLOWED_ORIGINS = [
    WEB_ORIGIN,
    "http://localhost:5174",
    "http://127.0.0.1:5174",
    "http://localhost:5173",
    "http://127.0.0.1:5173",
]

app = FastAPI(title="uni-q 2.0")
sio = socketio.AsyncServer(async_mode="asgi", cors_allowed_origins=ALLOWED_ORIGINS, cors_credentials=True)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, same_site="lax", https_only=NODE_ENV == "production")


async def broadcast_queue() -> None:
    await sio.emit("queue:update", get_live_queue())


@sio.event
async def connect(sid: str, environ: dict[str, Any], auth: Any = None) -> None:
    await sio.emit("queue:update", get_live_queue(), to=sid)


def require_manager(request: Request) -> int:
    manager_id = request.session.get("managerId") or request.session.get("advisorId")
    if not manager_id:
        raise HTTPException(status_code=401, detail={"error": "Не авторизован"})
    request.session["managerId"] = int(manager_id)
    request.session.pop("advisorId", None)
    return int(manager_id)


def require_admin(request: Request) -> int:
    admin_id = request.session.get("adminId")
    if not admin_id:
        raise HTTPException(status_code=401, detail={"error": "Нет доступа администратора"})
    return int(admin_id)


@app.exception_handler(HTTPException)
async def http_error(_request: Request, exc: HTTPException) -> JSONResponse:
    if isinstance(exc.detail, dict):
        return JSONResponse(exc.detail, status_code=exc.status_code)
    return JSONResponse({"error": str(exc.detail)}, status_code=exc.status_code)


@app.get("/api/session")
def api_session() -> dict[str, Any]:
    return get_queue_session()


@app.post("/api/session/start")
async def api_session_start(_: int = Depends(require_manager)) -> dict[str, Any]:
    execute("UPDATE queue_session SET is_active = 1 WHERE id = 1")
    await broadcast_queue()
    return get_queue_session()


@app.post("/api/session/stop")
async def api_session_stop(_: int = Depends(require_manager)) -> dict[str, Any]:
    execute("UPDATE queue_session SET is_active = 0 WHERE id = 1")
    await broadcast_queue()
    return get_queue_session()


@app.post("/api/registration/check")
async def registration_check(request: Request) -> dict[str, bool]:
    body = await request.json()
    return registration_open_for_student(
        {
            "school": body.get("school"),
            "specialty_code": body.get("specialtyCode"),
            "language_section": body.get("languageSection"),
            "course": body.get("course"),
            "study_duration_years": body.get("studyDurationYears"),
        }
    )


def split_name(full: str) -> dict[str, str]:
    parts = re.sub(r"\s+", " ", full.strip()).split(" ") if full.strip() else []
    if not parts:
        return {"firstName": "", "lastName": ""}
    return {"firstName": parts[0], "lastName": " ".join(parts[1:])}


def jwt_payload(token: str) -> dict[str, Any] | None:
    parts = token.split(".")
    if len(parts) < 2:
        return None
    try:
        pad = "=" * ((4 - len(parts[1]) % 4) % 4)
        raw = base64.urlsafe_b64decode((parts[1] + pad).encode())
        return json.loads(raw.decode("utf-8"))
    except Exception:
        return None


def microsoft_cfg() -> dict[str, str]:
    return {
        "tenant": os.getenv("MS_TENANT_ID", "common").strip() or "common",
        "client_id": os.getenv("MS_CLIENT_ID", "").strip(),
        "client_secret": os.getenv("MS_CLIENT_SECRET", "").strip(),
        "redirect_uri": os.getenv("MS_REDIRECT_URI", f"{WEB_ORIGIN}/api/auth/microsoft/callback").strip(),
    }


@app.get("/api/auth/microsoft/start")
def microsoft_start(request: Request) -> Response:
    cfg = microsoft_cfg()
    if not cfg["client_id"]:
        return PlainTextResponse("MS_CLIENT_ID is not configured", status_code=500)
    state = f"{int(datetime.now().timestamp() * 1000)}-{secrets.token_hex(8)}"
    request.session["msState"] = state
    params = urlencode(
        {
            "client_id": cfg["client_id"],
            "response_type": "code",
            "redirect_uri": cfg["redirect_uri"],
            "response_mode": "query",
            "scope": "openid profile email User.Read",
            "state": state,
        }
    )
    return RedirectResponse(f"https://login.microsoftonline.com/{cfg['tenant']}/oauth2/v2.0/authorize?{params}")


@app.get("/api/auth/microsoft/callback")
async def microsoft_callback(request: Request) -> Response:
    code = str(request.query_params.get("code") or "")
    state = str(request.query_params.get("state") or "")
    expected = str(request.session.pop("msState", "") or "")
    if not code:
        return RedirectResponse(f"{WEB_ORIGIN}/student?ms=error")
    if not state or state != expected:
        return RedirectResponse(f"{WEB_ORIGIN}/student?ms=state")
    cfg = microsoft_cfg()
    if not cfg["client_id"] or not cfg["client_secret"]:
        return RedirectResponse(f"{WEB_ORIGIN}/student?ms=cfg")
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            token = await client.post(
                f"https://login.microsoftonline.com/{cfg['tenant']}/oauth2/v2.0/token",
                data={
                    "client_id": cfg["client_id"],
                    "client_secret": cfg["client_secret"],
                    "grant_type": "authorization_code",
                    "code": code,
                    "redirect_uri": cfg["redirect_uri"],
                    "scope": "openid profile email User.Read",
                },
            )
        data = token.json()
        if token.status_code >= 400:
            return RedirectResponse(f"{WEB_ORIGIN}/student?ms=token")
        payload = jwt_payload(str(data.get("id_token") or "")) or {}
        first = str(payload.get("given_name") or "")
        last = str(payload.get("family_name") or "")
        display = str(payload.get("name") or "")
        if (not first or not last) and display:
            sp = split_name(display)
            first = first or sp["firstName"]
            last = last or sp["lastName"]
        request.session["student"] = {
            "oid": payload.get("oid") or payload.get("sub"),
            "email": payload.get("preferred_username") or payload.get("email"),
            "firstName": first or None,
            "lastName": last or None,
            "name": display or None,
        }
        return RedirectResponse(f"{WEB_ORIGIN}/student?ms=ok")
    except Exception:
        return RedirectResponse(f"{WEB_ORIGIN}/student?ms=error")


@app.post("/api/auth/microsoft/logout")
def microsoft_logout(request: Request) -> dict[str, bool]:
    request.session.pop("student", None)
    return {"ok": True}


@app.get("/api/student/me")
def student_me(request: Request) -> dict[str, Any]:
    return {"ok": True, "student": request.session.get("student")}


@app.post("/api/auth/login")
async def manager_login(request: Request) -> dict[str, Any]:
    body = await request.json()
    login, password = str(body.get("login") or ""), str(body.get("password") or "")
    if not login or not password:
        raise HTTPException(400, {"error": "Введите логин и пароль"})
    user = row("SELECT id, login, password_hash FROM advisors WHERE login = ?", (login,))
    if not user or not bcrypt_check(password, user.get("password_hash")):
        raise HTTPException(401, {"error": "Неверный логин или пароль"})
    request.session.pop("adminId", None)
    request.session["managerId"] = int(user["id"])
    return {"ok": True}


@app.post("/api/auth/logout")
def manager_logout(request: Request) -> dict[str, bool]:
    request.session.clear()
    return {"ok": True}


@app.post("/api/admin/login")
async def admin_login(request: Request) -> dict[str, Any]:
    body = await request.json()
    login, password = str(body.get("login") or ""), str(body.get("password") or "")
    if not login or not password:
        raise HTTPException(400, {"error": "Введите логин и пароль"})
    user = row("SELECT id, login, password_hash, name FROM admin_users WHERE login = ?", (login,))
    if not user or not bcrypt_check(password, user.get("password_hash")):
        raise HTTPException(401, {"error": "Неверный логин или пароль"})
    request.session.pop("managerId", None)
    request.session["adminId"] = int(user["id"])
    return {"ok": True, "id": user["id"], "login": user["login"], "name": user.get("name") or "Admin"}


@app.post("/api/admin/logout")
def admin_logout(request: Request) -> dict[str, bool]:
    request.session.pop("adminId", None)
    return {"ok": True}


@app.get("/api/admin/me")
def admin_me(admin_id: int = Depends(require_admin)) -> dict[str, Any]:
    r = row("SELECT id, login, name FROM admin_users WHERE id = ?", (admin_id,))
    if not r:
        raise HTTPException(401, {"error": "Нет доступа"})
    return {"id": r["id"], "login": r["login"], "name": r.get("name") or "Admin"}


@app.patch("/api/admin/me/password")
async def admin_password(request: Request, admin_id: int = Depends(require_admin)) -> dict[str, bool]:
    body = await request.json()
    current, new = str(body.get("currentPassword") or ""), str(body.get("newPassword") or "")
    if not current or not new:
        raise HTTPException(400, {"error": "Укажите текущий и новый пароль"})
    if len(new) < 6:
        raise HTTPException(400, {"error": "Новый пароль минимум 6 символов"})
    r = row("SELECT password_hash FROM admin_users WHERE id = ?", (admin_id,))
    if not r or not bcrypt_check(current, r.get("password_hash")):
        raise HTTPException(400, {"error": "Текущий пароль неверный"})
    execute("UPDATE admin_users SET password_hash = ? WHERE id = ?", (bcrypt_hash(new), admin_id))
    return {"ok": True}


@app.patch("/api/managers/me/password")
async def manager_password(request: Request, manager_id: int = Depends(require_manager)) -> dict[str, bool]:
    body = await request.json()
    current, new = str(body.get("currentPassword") or ""), str(body.get("newPassword") or "")
    if not current or not new:
        raise HTTPException(400, {"error": "Укажите текущий и новый пароль"})
    if len(new) < 6:
        raise HTTPException(400, {"error": "Новый пароль минимум 6 символов"})
    r = row("SELECT password_hash FROM advisors WHERE id = ?", (manager_id,))
    if not r or not bcrypt_check(current, r.get("password_hash")):
        raise HTTPException(400, {"error": "Текущий пароль неверный"})
    execute("UPDATE advisors SET password_hash = ? WHERE id = ?", (bcrypt_hash(new), manager_id))
    return {"ok": True}


@app.get("/api/admin/managers")
def admin_managers(request: Request, _: int = Depends(require_admin)) -> dict[str, Any]:
    day = parse_ymd(request.query_params.get("day")) or today_sql()
    return {
        "rows": rows(
            """SELECT a.id, a.name, a.faculty, a.department, a.desk_number, a.login,
                      a.assigned_schools_json, a.assigned_languages_json, a.assigned_courses_json,
                      a.assigned_specialties_json, a.assigned_study_years_json,
                      COALESCE(d.work_ms, 0) AS work_ms_today
               FROM advisors a
               LEFT JOIN advisor_work_daily d ON d.advisor_id = a.id AND d.day = ?
               ORDER BY a.id ASC""",
            (day,),
        )
    }


@app.post("/api/admin/managers")
async def admin_create_manager(request: Request, _: int = Depends(require_admin)) -> dict[str, Any]:
    body = await request.json()
    first, last = str(body.get("firstName") or "").strip(), str(body.get("lastName") or "").strip()
    login, password = str(body.get("login") or "").strip(), str(body.get("password") or "")
    if not first or not last:
        raise HTTPException(400, {"error": "Укажите имя и фамилию"})
    if not login:
        raise HTTPException(400, {"error": "Укажите логин"})
    if len(password) < 4:
        raise HTTPException(400, {"error": "Пароль не короче 4 символов"})
    if row("SELECT 1 FROM advisors WHERE login = ?", (login,)):
        raise HTTPException(409, {"error": "Логин уже занят"})
    cur = execute(
        """INSERT INTO advisors (
             name, faculty, department, desk_number, login, password_hash,
             assigned_schools_json, assigned_language, assigned_languages_json, assigned_courses_json,
             assigned_specialties_json, assigned_study_years_json, reception_open
           ) VALUES (?, NULL, NULL, NULL, ?, ?, '[]', NULL, NULL, '[1,2,3,4]', NULL, NULL, 1)""",
        (f"{first} {last}".strip(), login, bcrypt_hash(password)),
    )
    return {"ok": True, "id": cur.lastrowid}


@app.patch("/api/admin/managers/{manager_id}/desk")
async def admin_manager_desk(manager_id: int, request: Request, _: int = Depends(require_admin)) -> dict[str, bool]:
    body = await request.json()
    raw = body.get("window")
    window = None if raw in (None, "") else int(raw)
    if window is not None and not 1 <= window <= 6:
        raise HTTPException(400, {"error": "Окно должно быть от 1 до 6 или пусто"})
    if not row("SELECT 1 FROM advisors WHERE id = ?", (manager_id,)):
        raise HTTPException(404, {"error": "Сотрудник не найден"})
    if window is not None:
        execute("UPDATE advisors SET desk_number = NULL WHERE id != ? AND desk_number = ?", (manager_id, str(window)))
    execute("UPDATE advisors SET desk_number = ? WHERE id = ?", (None if window is None else str(window), manager_id))
    return {"ok": True}


@app.delete("/api/admin/managers/{manager_id}")
def admin_delete_manager(manager_id: int, _: int = Depends(require_admin)) -> dict[str, bool]:
    if not row("SELECT 1 FROM advisors WHERE id = ?", (manager_id,)):
        raise HTTPException(404, {"error": "Сотрудник не найден"})
    active = row("SELECT COUNT(*) AS c FROM tickets WHERE advisor_id = ? AND status IN ('WAITING','CALLED','IN_SERVICE')", (manager_id,))
    if int(active["c"]) > 0:
        raise HTTPException(409, {"error": "Нельзя удалить: у сотрудника есть активные талоны"})
    execute("UPDATE tickets SET advisor_id = NULL WHERE advisor_id = ?", (manager_id,))
    execute("DELETE FROM advisor_work_daily WHERE advisor_id = ?", (manager_id,))
    execute("DELETE FROM advisor_work_totals WHERE advisor_id = ?", (manager_id,))
    execute("DELETE FROM advisors WHERE id = ?", (manager_id,))
    return {"ok": True}


@app.get("/api/managers/me")
def manager_me(manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    return row(
        """SELECT a.id, a.name, a.faculty, a.department, a.desk_number, COALESCE(a.reception_open, 1) AS reception_open,
                  a.assigned_schools_json, a.assigned_language, a.assigned_languages_json, a.assigned_courses_json,
                  a.assigned_specialties_json, a.assigned_study_years_json, a.assigned_school_scopes_json,
                  COALESCE(w.total_ms, 0) AS total_work_ms
           FROM advisors a LEFT JOIN advisor_work_totals w ON w.advisor_id = a.id WHERE a.id = ?""",
        (manager_id,),
    ) or {}


@app.patch("/api/managers/me/reception")
async def manager_reception(request: Request, manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    body = await request.json()
    execute("UPDATE advisors SET reception_open = ? WHERE id = ?", (1 if body.get("open") else 0, manager_id))
    recompute_route_owners()
    await broadcast_queue()
    return manager_me(manager_id)


@app.patch("/api/managers/me/work-total")
async def manager_work_total(request: Request, manager_id: int = Depends(require_manager)) -> dict[str, bool]:
    body = await request.json()
    total = float(body.get("totalMs"))
    if total < 0:
        raise HTTPException(400, {"error": "Некорректное totalMs"})
    execute(
        """INSERT INTO advisor_work_totals (advisor_id, total_ms, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(advisor_id) DO UPDATE SET total_ms = MAX(total_ms, excluded.total_ms), updated_at = CURRENT_TIMESTAMP""",
        (manager_id, int(total)),
    )
    today = body.get("todayMs")
    if today is not None and float(today) >= 0:
        day = parse_ymd(body.get("day")) or today_sql()
        execute(
            """INSERT INTO advisor_work_daily (advisor_id, day, work_ms) VALUES (?, ?, ?)
               ON CONFLICT(advisor_id, day) DO UPDATE SET work_ms = MAX(work_ms, excluded.work_ms)""",
            (manager_id, day, int(float(today))),
        )
    return {"ok": True}


@app.patch("/api/managers/me/scope")
async def manager_scope(request: Request, manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    body = await request.json()
    schools = [str(x) for x in body.get("assigned_schools_json", []) if str(x)]
    if not schools:
        raise HTTPException(400, {"error": "Выберите хотя бы одну школу"})
    langs = [str(x).lower() for x in body.get("assigned_languages_json", []) if str(x)]
    courses = [int(x) for x in body.get("assigned_courses_json", [1, 2, 3, 4]) if 1 <= int(x) <= 4]
    specs = [str(x) for x in body.get("assigned_specialties_json", []) if str(x)]
    years = [parse_study_duration(x) for x in body.get("assigned_study_years_json", [])]
    years = [x for x in years if x is not None]
    school_scopes = parse_school_scopes(body.get("assigned_school_scopes_json"))
    execute(
        """UPDATE advisors SET assigned_schools_json = ?, assigned_languages_json = ?, assigned_courses_json = ?,
                  assigned_specialties_json = ?, assigned_study_years_json = ?, assigned_school_scopes_json = ?
           WHERE id = ?""",
        (
            json.dumps(schools, ensure_ascii=False),
            json.dumps(langs, ensure_ascii=False) if langs else None,
            json.dumps(courses or [1, 2, 3, 4]),
            json.dumps(specs, ensure_ascii=False) if specs else None,
            json.dumps(years) if years else None,
            json.dumps(school_scopes, ensure_ascii=False) if school_scopes else None,
            manager_id,
        ),
    )
    recompute_route_owners()
    await broadcast_queue()
    return manager_me(manager_id)


@app.get("/api/queue/live")
def queue_live() -> dict[str, Any]:
    return get_live_queue()


@app.get("/api/admin/queues/all")
def admin_all_queues(_: int = Depends(require_admin)) -> dict[str, Any]:
    advisors = rows("SELECT * FROM advisors ORDER BY id ASC")
    tickets = rows(
        """SELECT id, queue_number, status, student_first_name, student_last_name, school, specialty, specialty_code,
                  language_section, course, study_duration_years, route_advisor_id, advisor_id, advisor_name,
                  advisor_desk, preferred_slot_at, created_at
           FROM tickets WHERE status IN ('WAITING','CALLED','IN_SERVICE')
           ORDER BY CASE status WHEN 'WAITING' THEN 0 WHEN 'CALLED' THEN 1 ELSE 2 END, queue_number ASC"""
    )
    by_id = {int(a["id"]): a for a in advisors}
    out = []
    for t in tickets:
        owner_id = t.get("route_advisor_id") if t["status"] == "WAITING" else t.get("advisor_id")
        owner = by_id.get(int(owner_id)) if owner_id else None
        targets = [a for a in advisors if t["status"] == "WAITING" and ticket_matches_scope(t, advisor_scope(a))]
        t["formatted_number"] = format_queue_number(t["queue_number"])
        t["owner_manager_id"] = owner.get("id") if owner else (targets[0]["id"] if targets else None)
        t["owner_manager_name"] = (owner or {}).get("name") or t.get("advisor_name") or (", ".join(a["name"] for a in targets) if targets else None)
        t["owner_manager_desk"] = (owner or {}).get("desk_number") or t.get("advisor_desk")
        out.append(t)
    return {"rows": out}


@app.post("/api/tickets")
async def create_ticket(request: Request) -> dict[str, Any]:
    body = await request.json()
    first, last, school = str(body.get("firstName") or "").strip(), str(body.get("lastName") or "").strip(), str(body.get("school") or "").strip()
    if not first or not last or not school:
        raise HTTPException(400, {"error": "Заполните имя, фамилию и школу"})
    lang, course, spec_code = str(body.get("languageSection") or "").strip(), str(body.get("course") or "").strip(), str(body.get("specialtyCode") or "").strip()
    if not lang or not course or not spec_code:
        raise HTTPException(400, {"error": "Заполните все поля профиля"})
    duration = parse_study_duration(body.get("studyDurationYears"))
    if duration is None:
        raise HTTPException(400, {"error": "Выберите тип обучения"})
    reg = registration_open_for_student({"school": school, "specialty_code": spec_code, "language_section": lang, "course": course, "study_duration_years": duration})
    if not reg["matchesAny"]:
        raise HTTPException(409, {"error": "Нет линии приёма для указанных данных"})
    if not reg["open"]:
        raise HTTPException(409, {"error": "Запись по вашему направлению сейчас закрыта менеджером"})
    slot = None
    if body.get("preferredSlotAt"):
        dt = parse_dt(body.get("preferredSlotAt"))
        if dt is None:
            raise HTTPException(400, {"error": "Некорректное время брони"})
        if dt.timestamp() < datetime.now(timezone.utc).timestamp() - 60:
            raise HTTPException(400, {"error": "Выберите время в будущем"})
        slot = dt.isoformat()
    qn = next_queue_number()
    ticket_stub = {"school": school, "specialty_code": spec_code, "language_section": lang, "course": course, "study_duration_years": duration}
    route_id = pick_route_advisor_id(ticket_stub)
    cur = execute(
        """INSERT INTO tickets (queue_number, status, student_first_name, student_last_name, school, specialty, specialty_code,
                  language_section, course, study_duration_years, route_advisor_id, preferred_slot_at)
           VALUES (?, 'WAITING', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (qn, first, last, school, str(body.get("specialty") or "").strip(), spec_code, lang, course, duration, route_id, slot),
    )
    t = row("SELECT id, queue_number, status, school, specialty, specialty_code, language_section, course, study_duration_years, route_advisor_id, preferred_slot_at FROM tickets WHERE id = ?", (cur.lastrowid,))
    await broadcast_queue()
    return {**t, "formatted_number": format_queue_number(t["queue_number"]), "estimated_time": compute_estimated_minutes(t)}


@app.get("/api/tickets/{ticket_id}/status")
def ticket_status(ticket_id: int) -> dict[str, Any]:
    t = row(
        """SELECT t.*, CASE WHEN r.ticket_id IS NOT NULL THEN 1 ELSE 0 END AS has_review
           FROM tickets t LEFT JOIN ticket_reviews r ON r.ticket_id = t.id WHERE t.id = ?""",
        (ticket_id,),
    )
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    t["formatted_number"] = format_queue_number(t["queue_number"])
    t["estimated_time"] = compute_estimated_minutes(t) if t["status"] == "WAITING" else None
    return t


@app.post("/api/tickets/{ticket_id}/cancel")
async def cancel_ticket(ticket_id: int) -> dict[str, Any]:
    t = row("SELECT id, status FROM tickets WHERE id = ?", (ticket_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    if t["status"] != "WAITING":
        raise HTTPException(409, {"error": "Отмена доступна только для талонов в ожидании"})
    execute("UPDATE tickets SET status = 'CANCELLED', finished_at = CURRENT_TIMESTAMP WHERE id = ?", (ticket_id,))
    await broadcast_queue()
    return {"ok": True, "id": ticket_id}


def call_ticket(ticket_id: int, manager_id: int) -> None:
    a = row("SELECT id, name, desk_number, faculty, department FROM advisors WHERE id = ?", (manager_id,))
    execute(
        """UPDATE tickets SET status = 'CALLED', called_at = CURRENT_TIMESTAMP,
              advisor_id = ?, advisor_name = ?, advisor_desk = ?, advisor_faculty = ?, advisor_department = ?
           WHERE id = ?""",
        (manager_id, a.get("name"), a.get("desk_number"), a.get("faculty"), a.get("department"), ticket_id),
    )


@app.post("/api/tickets/call-next")
async def call_next(manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    a = row("SELECT * FROM advisors WHERE id = ?", (manager_id,))
    if not a:
        raise HTTPException(404, {"error": "Сотрудник не найден"})
    if int(a.get("reception_open") or 0) == 0:
        raise HTTPException(403, {"error": "Откройте запись студентов, чтобы вызывать"})
    waiting = rows("SELECT * FROM tickets WHERE status = 'WAITING' ORDER BY CASE WHEN preferred_slot_at IS NOT NULL THEN preferred_slot_at ELSE '9999-12-31' END ASC, queue_number ASC")
    next_ticket = next((t for t in waiting if ticket_matches_scope(t, advisor_scope(a)) and booking_callable_now(t.get("preferred_slot_at"))), None)
    if not next_ticket:
        raise HTTPException(404, {"error": "Нет студентов, доступных для вызова в вашей зоне. Если у всех есть бронь — дождитесь указанного времени или используйте «Позвать по брони»."})
    call_ticket(int(next_ticket["id"]), manager_id)
    await broadcast_queue()
    return {"ok": True, "ticketId": next_ticket["id"]}


@app.post("/api/tickets/{ticket_id}/call-booked")
async def call_booked(ticket_id: int, manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    t = row("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    a = row("SELECT * FROM advisors WHERE id = ?", (manager_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    if t["status"] != "WAITING":
        raise HTTPException(409, {"error": "Вызов доступен только для талона в ожидании"})
    if not t.get("preferred_slot_at"):
        raise HTTPException(400, {"error": "У талона нет брони времени — используйте «Вызвать следующего»"})
    if not booking_callable_now(t.get("preferred_slot_at")):
        raise HTTPException(409, {"error": "Нельзя вызвать раньше времени брони"})
    if not a or int(a.get("reception_open") or 0) == 0:
        raise HTTPException(403, {"error": "Откройте запись студентов, чтобы вызывать"})
    if not ticket_matches_scope(t, advisor_scope(a)):
        raise HTTPException(403, {"error": "Этот талон не относится к вашей зоне приёма"})
    call_ticket(ticket_id, manager_id)
    await broadcast_queue()
    return {"ok": True, "ticketId": ticket_id}


@app.post("/api/tickets/{ticket_id}/call-to-my-desk")
async def call_to_my_desk(ticket_id: int, manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    a = row("SELECT reception_open FROM advisors WHERE id = ?", (manager_id,))
    if not a or int(a["reception_open"]) == 0:
        raise HTTPException(403, {"error": "Откройте запись студентов, чтобы вызывать из очереди"})
    t = row("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    if t["status"] != "WAITING":
        raise HTTPException(409, {"error": "Вызов доступен только для талона в ожидании"})
    if not booking_callable_now(t.get("preferred_slot_at")):
        raise HTTPException(409, {"error": "Для этого талона ещё не наступило время брони"})
    call_ticket(ticket_id, manager_id)
    await broadcast_queue()
    return {"ok": True, "ticketId": ticket_id}


@app.patch("/api/tickets/{ticket_id}")
async def update_ticket(ticket_id: int, request: Request, _: int = Depends(require_manager)) -> dict[str, bool]:
    body = await request.json()
    t = row("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    for key in ["comment", "case_type", "case_subtype", "contact_type", "student_comment", "manager_attachment_name", "manager_attachment_data_url"]:
        if key in body:
            val = body.get(key)
            if key == "manager_attachment_data_url" and val and not str(val).startswith("data:"):
                raise HTTPException(400, {"error": "Некорректный файл"})
            execute(f"UPDATE tickets SET {key} = ? WHERE id = ?", (None if val is None or str(val).strip() == "" else str(val), ticket_id))
    if "send_email_requested" in body:
        execute("UPDATE tickets SET send_email_requested = ? WHERE id = ?", (1 if body.get("send_email_requested") else 0, ticket_id))
    if "status" in body:
        status = str(body.get("status") or "")
        if status not in VALID_STATUSES:
            raise HTTPException(400, {"error": "Неверный статус"})
        latest = row("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
        if status == "DONE":
            if str(latest.get("case_type") or "") not in {"ACADEMIC", "FINANCIAL", "STATEMENTS", "CERTIFICATES", "ONAY", "MILITARY_DEPT", "ACADEMIC_MOBILITY", "TECHNICAL"}:
                raise HTTPException(400, {"error": "Укажите категорию обращения"})
            if not str(latest.get("case_subtype") or "").strip():
                raise HTTPException(400, {"error": "Укажите подкатегорию обращения"})
            if str(latest.get("contact_type") or "") not in {"QUESTION", "CONSULTATION", "PROBLEM"}:
                raise HTTPException(400, {"error": "Укажите тип обращения"})
            wc = count_words(latest.get("comment"))
            if wc < 1:
                raise HTTPException(400, {"error": "Комментарий обязателен"})
            if wc > 300:
                raise HTTPException(400, {"error": "Комментарий не более 300 слов"})
        terminal = status in {"DONE", "MISSED", "CANCELLED"}
        was_terminal = t["status"] in {"DONE", "MISSED", "CANCELLED"}
        execute(
            """UPDATE tickets SET status = ?,
                  started_at = CASE WHEN ? = 1 THEN CURRENT_TIMESTAMP ELSE started_at END,
                  finished_at = CASE WHEN ? = 1 THEN CURRENT_TIMESTAMP ELSE finished_at END
               WHERE id = ?""",
            (status, 1 if status == "IN_SERVICE" else 0, 1 if terminal else 0, ticket_id),
        )
        if terminal and not was_terminal:
            logged = row("SELECT COUNT(*) AS c FROM ticket_visit_log WHERE ticket_id = ?", (ticket_id,))
            insert_visit_log(row("SELECT * FROM tickets WHERE id = ?", (ticket_id,)), int(logged["c"]) > 0)
    await broadcast_queue()
    return {"ok": True}


@app.post("/api/tickets/{ticket_id}/review")
async def review_ticket(ticket_id: int, request: Request) -> dict[str, bool]:
    body = await request.json()
    t = row("SELECT id, status FROM tickets WHERE id = ?", (ticket_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    if t["status"] != "DONE":
        raise HTTPException(409, {"error": "Отзыв доступен после завершения приёма"})
    if row("SELECT 1 FROM ticket_reviews WHERE ticket_id = ?", (ticket_id,)):
        raise HTTPException(409, {"error": "Отзыв уже отправлен"})
    stars = int(body.get("stars"))
    comment = str(body.get("comment") or "").strip()
    if stars < 1 or stars > 5:
        raise HTTPException(400, {"error": "Оцените от 1 до 5 звёзд"})
    if stars <= 3 and not comment:
        raise HTTPException(400, {"error": "Для оценки 3 и ниже комментарий обязателен"})
    execute("INSERT INTO ticket_reviews (ticket_id, stars, comment) VALUES (?, ?, ?)", (ticket_id, stars, comment or None))
    return {"ok": True}


@app.post("/api/tickets/{ticket_id}/missed-feedback")
async def missed_feedback(ticket_id: int, request: Request) -> dict[str, bool]:
    body = await request.json()
    t = row("SELECT id, status, missed_student_note FROM tickets WHERE id = ?", (ticket_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    if t["status"] != "MISSED":
        raise HTTPException(409, {"error": "Талон не в статусе пропуска"})
    if t.get("missed_student_note") is not None:
        raise HTTPException(409, {"error": "Уже отправлено"})
    execute("UPDATE tickets SET missed_student_note = ? WHERE id = ?", (str(body.get("reason") or "").strip()[:2000], ticket_id))
    return {"ok": True}


@app.post("/api/tickets/{ticket_id}/reopen")
async def reopen_ticket(ticket_id: int, request: Request, manager_id: int = Depends(require_manager)) -> dict[str, bool]:
    body = await request.json()
    action = str(body.get("action") or "")
    if action not in {"queue", "service", "comment"}:
        raise HTTPException(400, {"error": "Нужно action: queue | service | comment"})
    t = row("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    if not t:
        raise HTTPException(404, {"error": "Талон не найден"})
    if int(t.get("advisor_id") or 0) != manager_id:
        raise HTTPException(403, {"error": "Не ваш талон"})
    if t["status"] not in {"DONE", "MISSED"}:
        raise HTTPException(400, {"error": "Доступно только для завершённых визитов"})
    fin = parse_dt(t.get("finished_at"))
    if not fin or (datetime.now(timezone.utc).timestamp() - fin.timestamp()) / 60 > 60:
        raise HTTPException(400, {"error": "Прошло больше часа с завершения"})
    if action == "comment":
        execute("UPDATE tickets SET comment = ? WHERE id = ?", (str(body.get("comment") or "")[:12000], ticket_id))
    elif action == "queue":
        execute(
            """UPDATE tickets SET status = 'WAITING', queue_number = ?, route_advisor_id = ?, called_at = NULL,
                  started_at = NULL, finished_at = NULL, advisor_id = NULL, advisor_name = NULL, advisor_desk = NULL,
                  advisor_faculty = NULL, advisor_department = NULL WHERE id = ?""",
            (next_queue_number(), pick_route_advisor_id(t), ticket_id),
        )
    else:
        execute("UPDATE tickets SET status = 'IN_SERVICE', finished_at = NULL, started_at = COALESCE(started_at, CURRENT_TIMESTAMP) WHERE id = ?", (ticket_id,))
    await broadcast_queue()
    return {"ok": True}


@app.get("/api/managers/me/history")
def manager_history(request: Request, manager_id: int = Depends(require_manager)) -> dict[str, Any]:
    limit = min(max(int(request.query_params.get("limit", "200")), 1), 500)
    day = parse_ymd(request.query_params.get("date")) or today_sql()
    rs = rows(
        """SELECT l.id AS log_id, l.ticket_id AS id, l.*, t.student_comment, t.study_duration_years
           FROM ticket_visit_log l JOIN tickets t ON t.id = l.ticket_id
           WHERE l.advisor_id = ? AND date(l.finished_at, 'localtime') = ?
           ORDER BY l.finished_at DESC, l.id DESC LIMIT ?""",
        (manager_id, day, limit),
    )
    for r in rs:
        r["formatted_number"] = format_queue_number(r["queue_number"])
        r["queue_wait_minutes"] = minutes_between(r.get("created_at"), r.get("started_at"))
        r["desk_service_minutes"] = minutes_between(r.get("started_at"), r.get("finished_at"))
        r["total_minutes"] = minutes_between(r.get("created_at"), r.get("finished_at"))
        r["reopen_eligible"] = 1
    return {"rows": rs}


@app.post("/api/stats/event")
async def stats_event(request: Request) -> dict[str, bool]:
    body = await request.json()
    et = str(body.get("event_type") or "")
    if not et:
        raise HTTPException(400, {"error": "Нужен event_type"})
    execute("INSERT INTO stats_events (event_type, meta) VALUES (?, ?)", (et[:80], json.dumps(body.get("meta"), ensure_ascii=False) if "meta" in body else None))
    return {"ok": True}


@app.get("/api/admin/stats/summary")
def stats_summary(_: int = Depends(require_admin)) -> dict[str, Any]:
    events = rows("SELECT event_type, COUNT(*) AS count FROM stats_events GROUP BY event_type")
    reviews = row("SELECT COUNT(*) AS c FROM ticket_reviews")["c"]
    today = row("SELECT COUNT(*) AS c FROM tickets WHERE date(created_at) = date('now', 'localtime')")["c"]
    booked = row("SELECT COUNT(*) AS c FROM tickets WHERE preferred_slot_at IS NOT NULL AND status IN ('WAITING','CALLED','IN_SERVICE')")["c"]
    return {"events": events, "reviewsTotal": reviews, "ticketsToday": today, "bookedSlotsLive": booked}


@app.get("/api/admin/stats/faq-no-queue")
def faq_no_queue(request: Request, _: int = Depends(require_admin)) -> Response:
    from_d, to_d = parse_ymd(request.query_params.get("from")), parse_ymd(request.query_params.get("to"))
    sql = "SELECT date(created_at, 'localtime') AS day, COUNT(*) AS count FROM stats_events WHERE event_type = 'faq_no_queue'"
    params: list[Any] = []
    if from_d and to_d:
        sql += " AND date(created_at, 'localtime') >= ? AND date(created_at, 'localtime') <= ?"
        params += [from_d, to_d]
    sql += " GROUP BY day ORDER BY day ASC"
    rs = rows(sql, tuple(params))
    if str(request.query_params.get("format") or "json").lower() == "csv":
        return PlainTextResponse(csv_response(rs, ["day", "count"]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": 'attachment; filename="faq-no-queue.csv"'})
    return JSONResponse({"from": from_d, "to": to_d, "series": rs})


@app.get("/api/admin/stats/wait-times")
def wait_times(request: Request, _: int = Depends(require_admin)) -> Response:
    from_d, to_d = parse_ymd(request.query_params.get("from")), parse_ymd(request.query_params.get("to"))
    if not from_d or not to_d:
        raise HTTPException(400, {"error": "Укажите from и to в формате YYYY-MM-DD"})
    if from_d > to_d:
        raise HTTPException(400, {"error": "Дата «с» не может быть позже «по»"})
    status_filter = str(request.query_params.get("status") or "").strip().upper()
    school_q = str(request.query_params.get("school") or "").strip().lower()
    min_wait_raw = str(request.query_params.get("minWait") or "").strip()
    max_wait_raw = str(request.query_params.get("maxWait") or "").strip()
    min_wait = float(min_wait_raw) if min_wait_raw else None
    max_wait = float(max_wait_raw) if max_wait_raw else None
    rs = rows(
        """SELECT id AS ticket_id, queue_number, student_first_name, student_last_name, school, status,
                  created_at, called_at, started_at,
                  CASE WHEN called_at IS NOT NULL THEN (strftime('%s', called_at)-strftime('%s', created_at))/60.0
                       WHEN started_at IS NOT NULL THEN (strftime('%s', started_at)-strftime('%s', created_at))/60.0
                       ELSE NULL END AS wait_minutes
           FROM tickets WHERE date(created_at, 'localtime') >= ? AND date(created_at, 'localtime') <= ?""",
        (from_d, to_d),
    )
    rs = [r for r in rs if r.get("wait_minutes") is not None and float(r["wait_minutes"]) >= 0]
    if status_filter in VALID_STATUSES:
        rs = [r for r in rs if str(r.get("status") or "").upper() == status_filter]
    if school_q:
        rs = [r for r in rs if school_q in str(r.get("school") or "").lower()]
    if min_wait is not None:
        rs = [r for r in rs if float(r["wait_minutes"]) >= min_wait]
    if max_wait is not None:
        rs = [r for r in rs if float(r["wait_minutes"]) <= max_wait]
    waits = sorted(float(r["wait_minutes"]) for r in rs)
    avg = sum(waits) / len(waits) if waits else 0
    med = waits[len(waits)//2] if waits else 0
    for r in rs:
        r["formatted_number"] = format_queue_number(r["queue_number"])
        r["wait_minutes"] = round(float(r["wait_minutes"]), 2)
    if str(request.query_params.get("format") or "json").lower() == "csv":
        return PlainTextResponse(csv_response(rs, ["ticket_id", "queue_number", "wait_minutes", "status", "created_at", "called_at", "started_at", "student_last_name", "student_first_name", "school"]), media_type="text/csv; charset=utf-8")
    return JSONResponse({"from": from_d, "to": to_d, "summary": {"count": len(rs), "avgMin": avg, "medianMin": med}, "rows": rs})


@app.get("/api/admin/stats/schools-served")
def schools_served(request: Request, _: int = Depends(require_admin)) -> Response:
    from_d, to_d = parse_ymd(request.query_params.get("from")), parse_ymd(request.query_params.get("to"))
    if not from_d or not to_d:
        raise HTTPException(400, {"error": "Укажите from и to в формате YYYY-MM-DD"})
    rs = rows("SELECT school, COUNT(*) AS count FROM ticket_visit_log WHERE status = 'DONE' AND date(finished_at, 'localtime') >= ? AND date(finished_at, 'localtime') <= ? GROUP BY school ORDER BY count DESC", (from_d, to_d))
    if str(request.query_params.get("format") or "json").lower() == "csv":
        return PlainTextResponse(csv_response(rs, ["school", "count"]), media_type="text/csv; charset=utf-8")
    return JSONResponse({"from": from_d, "to": to_d, "rows": rs})


@app.get("/api/admin/stats/bookings")
def bookings(request: Request, _: int = Depends(require_admin)) -> Response:
    from_d, to_d = parse_ymd(request.query_params.get("from")), parse_ymd(request.query_params.get("to"))
    if not from_d or not to_d:
        raise HTTPException(400, {"error": "Укажите from и to в формате YYYY-MM-DD"})
    if from_d > to_d:
        raise HTTPException(400, {"error": "Дата «с» не может быть позже «по»"})
    status_filter = str(request.query_params.get("status") or "").strip().upper()
    school_q = str(request.query_params.get("school") or "").strip().lower()
    manager_raw = str(request.query_params.get("managerId") or "").strip()
    manager_id = int(manager_raw) if manager_raw.isdigit() and int(manager_raw) > 0 else None
    rs = rows(
        """SELECT t.id AS ticket_id, t.queue_number, t.student_first_name, t.student_last_name, t.school, t.specialty,
                  t.preferred_slot_at, t.status, t.created_at, COALESCE(t.advisor_name, a.name) AS advisor_name,
                  COALESCE(t.advisor_desk, a.desk_number) AS advisor_desk, t.route_advisor_id, t.advisor_id
           FROM tickets t LEFT JOIN advisors a ON a.id = t.route_advisor_id
           WHERE t.preferred_slot_at IS NOT NULL AND date(t.preferred_slot_at, 'localtime') >= ? AND date(t.preferred_slot_at, 'localtime') <= ?
           ORDER BY t.preferred_slot_at ASC, t.id ASC""",
        (from_d, to_d),
    )
    if status_filter in VALID_STATUSES:
        rs = [r for r in rs if str(r.get("status") or "").upper() == status_filter]
    if school_q:
        rs = [r for r in rs if school_q in str(r.get("school") or "").lower()]
    if manager_id is not None:
        rs = [
            r for r in rs
            if int(r.get("route_advisor_id") or 0) == manager_id
            or int(r.get("advisor_id") or 0) == manager_id
        ]
    for r in rs:
        r["formatted_number"] = format_queue_number(r["queue_number"])
    if str(request.query_params.get("format") or "json").lower() == "csv":
        return PlainTextResponse(csv_response(rs, ["ticket_id", "queue_number", "preferred_slot_at", "status", "created_at", "student_last_name", "student_first_name", "school", "specialty", "advisor_name", "advisor_desk"]), media_type="text/csv; charset=utf-8")
    return JSONResponse({"from": from_d, "to": to_d, "rows": rs})


@app.get("/api/admin/stats/load")
def load_stats(request: Request, _: int = Depends(require_admin)) -> dict[str, Any]:
    date = parse_ymd(request.query_params.get("date"))
    if not date:
        raise HTTPException(400, {"error": "Укажите дату в формате YYYY-MM-DD"})
    year, month = int(date[:4]), int(date[5:7])
    status_filter = str(request.query_params.get("status") or "").strip().upper()
    status = status_filter if status_filter in VALID_STATUSES else ""
    manager_raw = str(request.query_params.get("managerId") or "").strip()
    manager_id = int(manager_raw) if manager_raw.isdigit() and int(manager_raw) > 0 else None
    extra = " AND (? = '' OR status = ?) AND (? IS NULL OR COALESCE(route_advisor_id, advisor_id) = ?)"
    daily = []
    for d in range(1, 32):
        dd = f"{year:04d}-{month:02d}-{d:02d}"
        daily.append({
            "day": d,
            "registrations": row(
                f"SELECT COUNT(*) AS c FROM tickets WHERE date(created_at, 'localtime') = ?{extra}",
                (dd, status, status, manager_id, manager_id),
            )["c"],
            "calls": row(
                f"SELECT COUNT(*) AS c FROM tickets WHERE date(called_at, 'localtime') = ?{extra}",
                (dd, status, status, manager_id, manager_id),
            )["c"],
        })
    monthly = []
    for m in range(1, 13):
        monthly.append({
            "month": m,
            "registrations": row(
                f"SELECT COUNT(*) AS c FROM tickets WHERE strftime('%Y', created_at, 'localtime') = ? AND strftime('%m', created_at, 'localtime') = ?{extra}",
                (str(year), f"{m:02d}", status, status, manager_id, manager_id),
            )["c"],
            "calls": row(
                f"SELECT COUNT(*) AS c FROM tickets WHERE called_at IS NOT NULL AND strftime('%Y', called_at, 'localtime') = ? AND strftime('%m', called_at, 'localtime') = ?{extra}",
                (str(year), f"{m:02d}", status, status, manager_id, manager_id),
            )["c"],
        })
    return {"year": year, "month": month, "daily": daily, "monthly": monthly}


@app.get("/api/admin/visits/history")
def visits_history(request: Request, _: int = Depends(require_admin)) -> Response:
    from_d, to_d = parse_ymd(request.query_params.get("from")), parse_ymd(request.query_params.get("to"))
    if not from_d or not to_d:
        raise HTTPException(400, {"error": "Укажите from и to в формате YYYY-MM-DD"})
    if from_d > to_d:
        raise HTTPException(400, {"error": "Дата «с» не может быть позже «по»"})
    status_filter = str(request.query_params.get("status") or "").strip().upper()
    school_q = str(request.query_params.get("school") or "").strip().lower()
    student_q = str(request.query_params.get("studentName") or "").strip().lower()
    rs = rows(
        """SELECT l.*, l.ticket_id, t.student_comment, t.study_duration_years
           FROM ticket_visit_log l LEFT JOIN tickets t ON t.id = l.ticket_id
           WHERE date(l.finished_at, 'localtime') >= ? AND date(l.finished_at, 'localtime') <= ?
           ORDER BY l.finished_at DESC, l.id DESC""",
        (from_d, to_d),
    )
    if status_filter in {"DONE", "MISSED", "CANCELLED"}:
        rs = [r for r in rs if str(r.get("status") or "").upper() == status_filter]
    if school_q:
        rs = [r for r in rs if school_q in str(r.get("school") or "").lower()]
    if student_q:
        rs = [
            r for r in rs
            if student_q in f"{str(r.get('student_last_name') or '').strip()} {str(r.get('student_first_name') or '').strip()}".lower()
        ]
    for r in rs:
        r["formatted_number"] = format_queue_number(r["queue_number"])
        r["queue_wait_minutes"] = minutes_between(r.get("created_at"), r.get("started_at"))
        r["desk_service_minutes"] = minutes_between(r.get("started_at"), r.get("finished_at"))
        r["total_minutes"] = minutes_between(r.get("created_at"), r.get("finished_at"))
    if str(request.query_params.get("format") or "json").lower() == "csv":
        return PlainTextResponse(csv_response(rs, ["ticket_id", "queue_number", "status", "student_last_name", "student_first_name", "school", "specialty", "language_section", "course", "study_duration_years", "advisor_name", "advisor_desk", "case_type", "comment", "student_comment", "called_at", "started_at", "finished_at"]), media_type="text/csv; charset=utf-8")
    return JSONResponse({"from": from_d, "to": to_d, "rows": rs})


@app.get("/api/admin/stats/reviews")
def reviews(request: Request, _: int = Depends(require_admin)) -> Response:
    from_d, to_d = parse_ymd(request.query_params.get("from")), parse_ymd(request.query_params.get("to"))
    if not from_d or not to_d:
        raise HTTPException(400, {"error": "Укажите from и to в формате YYYY-MM-DD"})
    if from_d > to_d:
        raise HTTPException(400, {"error": "Дата «с» не может быть позже «по»"})
    stars_raw = str(request.query_params.get("stars") or "").strip()
    stars = int(stars_raw) if stars_raw.isdigit() and 1 <= int(stars_raw) <= 5 else None
    school_q = str(request.query_params.get("school") or "").strip().lower()
    rs = rows(
        """SELECT r.ticket_id, r.stars, r.comment AS review_comment, r.created_at AS review_at,
                  t.queue_number, t.student_first_name, t.student_last_name, t.advisor_name, t.advisor_desk,
                  t.school, t.specialty, t.finished_at AS visit_finished_at
           FROM ticket_reviews r JOIN tickets t ON t.id = r.ticket_id
           WHERE date(r.created_at, 'localtime') >= ? AND date(r.created_at, 'localtime') <= ?
           ORDER BY r.created_at DESC, r.ticket_id DESC""",
        (from_d, to_d),
    )
    if stars is not None:
        rs = [r for r in rs if int(r.get("stars") or 0) == stars]
    if school_q:
        rs = [r for r in rs if school_q in str(r.get("school") or "").lower()]
    for r in rs:
        r["formatted_number"] = format_queue_number(r["queue_number"])
    if str(request.query_params.get("format") or "json").lower() == "csv":
        return PlainTextResponse(csv_response(rs, ["ticket_id", "queue_number", "review_at", "stars", "student_last_name", "student_first_name", "advisor_name", "advisor_desk", "school", "specialty", "visit_finished_at", "review_comment"]), media_type="text/csv; charset=utf-8")
    return JSONResponse({"from": from_d, "to": to_d, "rows": rs})


def normalize_kb(text: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", text.lower().replace("ё", "е"), flags=re.U)).strip()


SITE_STATIC_QA = [
    (re.compile(r"занят.*очередь|встат.*очередь|получить.*талон|взять.*талон|записат.*очередь", re.I), "Чтобы встать в очередь: нажмите **«Продолжить без входа»**, заполните форму студента и нажмите **«Получить талон»**. Когда вас вызовут, подойдите к указанному окну."),
    (re.compile(r"бронир|слот|временн.*запись", re.I), "Чтобы забронировать время, откройте Панель студента, нажмите **«Забронировать время»** и выберите слот на сегодня."),
    (re.compile(r"схем.*офис|где.*окна|где.*прием", re.I), "Схема офиса доступна по кнопке **«Схема офиса»**. Приёмная находится: **главный корпус, 2-й этаж, зона у лифтов**."),
]


@app.post("/api/student/chat")
async def student_chat(request: Request) -> dict[str, Any]:
    body = await request.json()
    messages = body.get("messages")
    if not isinstance(messages, list) or not messages:
        raise HTTPException(400, {"error": "chat_invalid"})
    last = next((str(m.get("content") or "").strip() for m in reversed(messages) if isinstance(m, dict) and m.get("role") == "user"), "")
    if not last:
        raise HTTPException(400, {"error": "chat_invalid"})
    for rx, answer in SITE_STATIC_QA:
        if rx.search(last):
            return {"reply": answer, "source": "static_site", "kbQuestionNorm": None}
    # Lightweight KB fallback: enough for local operation without Node/xlsx.
    try:
        import openpyxl

        if CHAT_KB_XLSX_PATH.exists():
            wb = openpyxl.load_workbook(CHAT_KB_XLSX_PATH, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            qn = set(normalize_kb(last).split())
            best: tuple[int, str, str] | None = None
            for r in ws.iter_rows(min_row=2, values_only=True):
                q, a = str(r[1] or ""), str(r[2] or "")
                if not q or not a:
                    continue
                score = len(qn & set(normalize_kb(q).split()))
                if score and (best is None or score > best[0]):
                    best = (score, q, a)
            if best and best[0] >= 2:
                return {"reply": best[2].strip(), "source": "local_kb_best", "kbQuestionNorm": normalize_kb(best[1])}
    except Exception:
        pass
    if not UNIQ_NVIDIA_API_KEY:
        return {"reply": "По этому вопросу информации нет. Обратитесь, пожалуйста, в Студенческий сервисный центр (SSC uni-q, каб. 123).", "source": "no_kb_match", "kbQuestionNorm": None}
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(
            f"{UNIQ_NVIDIA_API_BASE}/chat/completions",
            headers={"Authorization": f"Bearer {UNIQ_NVIDIA_API_KEY}"},
            json={"model": UNIQ_NVIDIA_CHAT_MODEL, "messages": messages[-12:], "max_tokens": 1024, "temperature": 0.1},
        )
    if r.status_code >= 400:
        raise HTTPException(502, {"error": "chat_upstream"})
    data = r.json()
    reply = data.get("choices", [{}])[0].get("message", {}).get("content")
    return {"reply": str(reply or "").strip(), "source": "nvidia", "kbQuestionNorm": None}


@app.post("/api/student/chat/feedback")
async def chat_feedback(request: Request) -> dict[str, bool]:
    body = await request.json()
    helpful = int(body.get("helpful") or 0)
    if helpful not in {1, -1}:
        raise HTTPException(400, {"error": "feedback_invalid"})
    user_q = str(body.get("userQuestion") or "").strip()[:6000]
    execute(
        """INSERT INTO chat_feedback (user_question, user_question_norm, answer_text, kb_question_norm, source, helpful)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (
            user_q or None,
            normalize_kb(user_q) if user_q else None,
            str(body.get("answer") or "").strip()[:12000] or None,
            str(body.get("kbQuestionNorm") or "").strip()[:600] or None,
            str(body.get("source") or "").strip()[:120] or None,
            helpful,
        ),
    )
    return {"ok": True}


if PUBLIC_DIR.exists():
    app.mount("/public", StaticFiles(directory=PUBLIC_DIR), name="public")
if FLAPPY_DIR.exists():
    app.mount("/flappy-bird", StaticFiles(directory=FLAPPY_DIR, html=True), name="flappy-bird")
if DIST_DIR.exists():
    app.mount("/assets", StaticFiles(directory=DIST_DIR / "assets"), name="assets")


@app.get("/{path:path}")
def spa_fallback(path: str) -> Response:
    public_target = PUBLIC_DIR / path
    if path and public_target.exists() and public_target.is_file():
        return FileResponse(public_target)
    target = DIST_DIR / path
    if path and target.exists() and target.is_file():
        return FileResponse(target)
    index = DIST_DIR / "index.html"
    if index.exists():
        return FileResponse(index)
    return PlainTextResponse(f"uni-q 2.0 Python server. SQLite: {SQLITE_PATH}", status_code=200)


socket_app = socketio.ASGIApp(sio, other_asgi_app=app)
